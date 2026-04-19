"""
attendance_manager.py
---------------------
Handles all attendance logging:
  - Excel file creation / loading (attendance.xlsx)
  - Duplicate entry prevention (same ID logged only once per session)
  - Status assignment: "Present" before 09:00, "Late" after
  - Unknown face image saving
"""

import os
import shutil
import datetime
import logging
import threading

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

ATTENDANCE_FILE = "attendance.xlsx"
UNKNOWN_DIR = "unknown_faces"
CUTOFF_TIME = datetime.time(9, 0, 0)  # 09:00:00

# Column definitions
COLUMNS = ["Name", "ID", "Time", "Date", "Status"]

# Thread lock for file I/O safety
_lock = threading.Lock()

# In-memory set of IDs already logged this session
_logged_ids: set = set()


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _create_excel_if_missing():
    """Create attendance.xlsx with styled headers if it doesn't exist."""
    if os.path.exists(ATTENDANCE_FILE):
        return

    df = pd.DataFrame(columns=COLUMNS)
    df.to_excel(ATTENDANCE_FILE, index=False, engine="openpyxl")

    # Apply header styling
    wb = load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="E94560", size=12)
    border = Border(
        bottom=Side(style="thin", color="E94560")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Set column widths
    col_widths = {"A": 20, "B": 12, "C": 14, "D": 14, "E": 12}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(ATTENDANCE_FILE)
    logger.info(f"Created new attendance file: {ATTENDANCE_FILE}")


def _load_logged_ids():
    """Populate _logged_ids from existing Excel data (avoid re-logging on restart)."""
    global _logged_ids
    _create_excel_if_missing()
    try:
        df = pd.read_excel(ATTENDANCE_FILE, engine="openpyxl")
        today = datetime.date.today().strftime("%Y-%m-%d")
        today_df = df[df["Date"].astype(str) == today]
        _logged_ids = set(today_df["ID"].astype(str).tolist())
        logger.info(f"Pre-loaded logged IDs for today: {_logged_ids}")
    except Exception as e:
        logger.error(f"Error loading logged IDs: {e}")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def initialize():
    """Call once at startup."""
    os.makedirs(UNKNOWN_DIR, exist_ok=True)
    _create_excel_if_missing()
    _load_logged_ids()


def mark_attendance(name: str, person_id: str) -> bool:
    """
    Log attendance for a recognized person.

    Returns True if a new record was written, False if duplicate.
    """
    with _lock:
        if person_id in _logged_ids:
            return False  # Already logged today

        now = datetime.datetime.now()
        time_str = now.strftime("%H:%M:%S")
        date_str = now.strftime("%Y-%m-%d")
        status = "Present" if now.time() < CUTOFF_TIME else "Late"

        record = {
            "Name": name,
            "ID": person_id,
            "Time": time_str,
            "Date": date_str,
            "Status": status,
        }

        try:
            df = pd.read_excel(ATTENDANCE_FILE, engine="openpyxl")
            new_row = pd.DataFrame([record])
            df = pd.concat([df, new_row], ignore_index=True)
            df.to_excel(ATTENDANCE_FILE, index=False, engine="openpyxl")

            # Re-apply header style (pandas overwrites it)
            _restyle_header()

            _logged_ids.add(person_id)
            logger.info(f"[ATTENDANCE] {name} (ID: {person_id}) — {status} at {time_str}")
            return True

        except Exception as e:
            logger.error(f"Error writing attendance: {e}")
            return False


def save_unknown_face(frame: np.ndarray, face_loc: tuple):
    """Crop and save an unknown face image with a timestamp filename."""
    top, right, bottom, left = face_loc
    # Add padding
    pad = 20
    h, w = frame.shape[:2]
    top = max(0, top - pad)
    left = max(0, left - pad)
    bottom = min(h, bottom + pad)
    right = min(w, right + pad)

    face_img = frame[top:bottom, left:right]
    if face_img.size == 0:
        return

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    filename = os.path.join(UNKNOWN_DIR, f"unknown_{timestamp}.jpg")
    cv2.imwrite(filename, face_img)
    logger.warning(f"Unknown face saved: {filename}")


def get_attendance_records() -> list[dict]:
    """
    Read the Excel file and return all records as a list of dicts.
    Sorted by Date + Time descending (most recent first).
    """
    _create_excel_if_missing()
    try:
        df = pd.read_excel(ATTENDANCE_FILE, engine="openpyxl")
        df = df.fillna("")
        # Sort most-recent first
        df = df.sort_values(by=["Date", "Time"], ascending=False)
        return df.to_dict(orient="records")
    except Exception as e:
        logger.error(f"Error reading attendance: {e}")
        return []


def copy_known_face_to_static(src_path: str, static_dir: str, person_id: str) -> str:
    """
    Copy a known-face image to static/images/ for web serving.
    Returns the web-accessible path like 'static/images/101.jpg'
    """
    os.makedirs(static_dir, exist_ok=True)
    ext = os.path.splitext(src_path)[1]
    dest_name = f"face_{person_id}{ext}"
    dest_path = os.path.join(static_dir, dest_name)
    if not os.path.exists(dest_path):
        shutil.copy2(src_path, dest_path)
    return f"images/{dest_name}"


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _restyle_header():
    """Re-apply header styles after pandas overwrites them."""
    try:
        wb = load_workbook(ATTENDANCE_FILE)
        ws = wb.active
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(bold=True, color="E94560", size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        col_widths = {"A": 20, "B": 12, "C": 14, "D": 14, "E": 12}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        wb.save(ATTENDANCE_FILE)
    except Exception:
        pass
